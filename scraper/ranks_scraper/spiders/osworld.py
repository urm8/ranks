from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

from ranks_scraper.parsing.normalize import attach_normalized_scores, ranking_row
from ranks_scraper.spiders.base import SourceSpider


VERIFIED_XLSX_URL = "https://os-world.github.io/static/data/osworld_verified_results.xlsx"
BROWSE_URL = "https://os-world.github.io/"


def _is_yes(val: Any) -> bool:
    if val is None:
        return False
    return str(val).strip().lower() in {"yes", "true", "1", "y"}


def _parse_rate(val: Any) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    text = str(val).strip()
    if text in {"🚧", "-", "n/a", "N/A"}:
        return None
    try:
        return float(text.replace("%", "").strip())
    except ValueError:
        return None


def _excel_date(val: Any) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    # Excel serial date (rare with data_only=True + datetime cells)
    if isinstance(val, (int, float)) and val > 20000:
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(val).date().isoformat()
        except Exception:  # noqa: BLE001
            return None
    text = str(val).strip()
    return text or None


class OsworldSpider(SourceSpider):
    """OSWorld verified leaderboard from the site's XLSX (not the empty HTML shell)."""

    name = "osworld"
    source_title = "OSWorld"
    start_urls = [VERIFIED_XLSX_URL]
    browse_url = BROWSE_URL
    kind = "benchmark"
    engine = "http"
    # Default public tab on os-world.github.io is "Foundation E2E GUI".
    foundation_only: bool = True

    def parse_osworld_xlsx(self, payload: bytes) -> list[dict]:
        wb = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            return []
        idx = {name: i for i, name in enumerate(header) if name}

        def cell(row: tuple[Any, ...], name: str) -> Any:
            i = idx.get(name)
            if i is None or i >= len(row):
                return None
            return row[i]

        a11y_key = next((h for h in header if h.lower() == "additional a11y tree used"), None)
        tool_key = next(
            (
                h
                for h in header
                if h.lower() in {"additional coding-based action", "additional tool used"}
            ),
            None,
        )

        groups: dict[tuple[str, str], dict[str, Any]] = {}
        for row in rows_iter:
            if not row:
                continue
            model = cell(row, "Model")
            if not model or not str(model).strip():
                continue
            rate = _parse_rate(cell(row, "Success rate"))
            if rate is None:
                continue
            model_name = str(model).strip()
            steps_raw = cell(row, "Max steps")
            steps = str(int(steps_raw)) if isinstance(steps_raw, (int, float)) else str(steps_raw or "").strip()
            key = (model_name, steps or "Unknown")
            group = groups.get(key)
            if group is None:
                group = {
                    "model": model_name,
                    "steps": steps,
                    "approach": cell(row, "Approach type"),
                    "institution": cell(row, "Institution"),
                    "date": _excel_date(cell(row, "Date")),
                    "rates": [],
                    "has_a11y": False,
                    "has_tool": False,
                }
                groups[key] = group
            group["rates"].append(rate)
            if a11y_key and _is_yes(cell(row, a11y_key)):
                group["has_a11y"] = True
            if tool_key and _is_yes(cell(row, tool_key)):
                group["has_tool"] = True
            date = _excel_date(cell(row, "Date"))
            if date and (not group["date"] or date > group["date"]):
                group["date"] = date

        aggregated: list[dict[str, Any]] = []
        for group in groups.values():
            rates = group["rates"]
            if not rates:
                continue
            mean = sum(rates) / len(rates)
            approach = str(group.get("approach") or "").strip()
            is_foundation = approach in {"General model", "Specialized model"} and not group["has_a11y"] and not group["has_tool"]
            aggregated.append(
                {
                    **group,
                    "score": mean,
                    "runs": len(rates),
                    "is_foundation": is_foundation,
                }
            )

        if self.foundation_only:
            aggregated = [g for g in aggregated if g["is_foundation"]]

        aggregated.sort(key=lambda g: (-float(g["score"]), g["model"], g["steps"]))

        rows_out: list[dict] = []
        seen: set[str] = set()
        for group in aggregated:
            steps = group["steps"]
            base_id = self.guess_model_id(group["model"], creator=str(group.get("institution") or "") or None)
            mid = f"{base_id}-s{steps}" if steps and steps != "Unknown" else base_id
            if not mid or mid in seen:
                continue
            seen.add(mid)
            score = float(group["score"])
            runs = int(group["runs"])
            note_bits = [f"{score:g}% success"]
            if steps and steps != "Unknown":
                note_bits.append(f"max steps {steps}")
            if runs > 1:
                note_bits.append(f"{runs} runs")
            if group.get("date"):
                note_bits.append(str(group["date"]))
            rank = len(rows_out) + 1
            row = ranking_row(
                rank=rank,
                model_id=mid,
                score=f"{score:g}%",
                note="; ".join(note_bits),
                normalized_score=score,
            )
            row["model"] = group["model"]
            if group.get("institution"):
                row["vendor_hint"] = str(group["institution"])
            rows_out.append(row)
        return attach_normalized_scores(rows_out)

    def parse_response(self, response):
        payload = response.body or b""
        rankings = self.parse_osworld_xlsx(payload)
        models = self.models_from_rankings(rankings)
        return {
            "rankings": rankings,
            "models": models,
            "mentions": [r["model_id"] for r in rankings],
        }


if __name__ == "__main__":
    from pathlib import Path

    sample = Path(__file__).with_name("osworld_sample.xlsx")
    legacy = Path(__file__).with_name("osworld_sample.html")
    spider = OsworldSpider()
    if sample.exists():
        rankings = spider.parse_osworld_xlsx(sample.read_bytes())
        print(sample.name, "parsed", len(rankings))
        for row in rankings[:10]:
            print(row["rank"], row.get("model"), row.get("score"), row.get("note"))
    elif legacy.exists():
        print(legacy.name, "is HTML-only; OSWorld rankings live in verified XLSX")
    else:
        print("no sample")
